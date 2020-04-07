

import xlrd,openpyxl,string,os
from openpyxl import styles
from openpyxl.chart import PieChart,ProjectedPieChart,Reference


class caseExcel:
    '''
    读取excel版用例的方法
    '''
    def __init__(self):
        self.excelPath = ''         #excel路径
        self.wb = ''        #打开文件后的操作
        self.sh = ''
        self.merge = ''         #合并单元格 list [(1,2,3,4),(2,3,4,5)]   起始行,结束行,起始列,结束列
        self.rowsValue = ''         #首行内容 list [da,da,da]
        self.nrows = 0             #最大有效行数 int
        self.ncols = 0              #最大有效列 int
        self.caseList = []          #一个sheet的case
        self.excelDic = {}          #一个excel的case  {sheet名字:列表case}


    def sheetMerge(self,rows,cols):
        '''
        根据传入的坐标,分析存在那一部分单元格内,返回它的首格坐标
        :param rows: 行 int
        :param cols: 列 int
        :return: 本合并单元格的开始行,开始列
        '''
        for array in self.merge:  # 循环拿到一组数据
            starow = int(array[0])
            endrow = int(array[1])
            stacol = int(array[2])
            endcol = int(array[3])
            for row in range(starow,endrow):
                if row == rows:
                    for col in range(stacol,endcol):
                        if col == cols:
                            return starow,stacol
        return None


    def dispose(self):
        '''
        获取sheet页内容
        :return:
        '''

        for row in range(1,self.nrows):
            row_listValue = []
            for col in range(self.ncols):
                value = self.sh.cell_value(rowx=row,colx=col)   #获取当个单元格的内容
                if not value:                                   #如果是合并单元格,取出来值为None
                    dinate = self.sheetMerge(row,col)
                    if dinate:
                        value = self.sh.cell_value(rowx=dinate[0],colx=dinate[1])
                row_listValue.append(value)                     #循环完成一列 代表一行数据取到添加到列表中
            rowDic = dict(zip(self.rowsValue,row_listValue))    #一行请求 dic 添加到列表中
            if rowDic['id']:
                self.caseList.append(rowDic)
        return

    def status(self,path):
        '''
        启动方法调用函数
        :return: 整理到excelDic里面 所有用例
        '''
        self.excelPath = path
        wb = xlrd.open_workbook(self.excelPath)
        for name in wb.sheet_names():
            self.sh = wb.sheet_by_name(name)
            self.merge = self.sh.merged_cells
            self.nrows = self.sh.nrows
            self.ncols = self.sh.ncols
            self.rowsValue = self.sh.row_values(0,end_colx=self.ncols)
            self.dispose()              #执行此方法获取整个sheet的case 整理成列表
            self.excelDic[name] = self.caseList
            self.caseList = []


class excelReport:
    '''
    处理测试报告,生成测试报告首页的方法
    '''
    Gold = "FFFFFF00"
    LimeGreen = "FF47FF00"
    red = "FFFF3F00"
    DoderBlue = "FF0016FF"
    white = "FFFFFFFF"
    Orange = "FFCD8500"
    Green = "FF008B00"

    def __init__(self):
        self.wb = ''
        self.ws = ''
        self.font = ''              #字体设置
        self.fill = ''              #单元格设置
        self.border = ''            #定义框架属性
        self.Alignment = ''         #定义文本样式,居中还是对其
        self.letter = letter()
        self.pie = ''
        self.caseTrue = 0
        self.caseFalse = 0

    def stylestFout(self,size,bold=None,color="FF010701"):
        '''
        单元格字体的设置
        :param size: 字体大小 int
        :param bold: 是否加粗,默认为None,需要传 True
        :param color: 颜色默认黑色,需要传入其他
        :return: 无
        '''
        self.font = styles.Font(name=u'微软雅黑',size=size,bold=bold,color=color)
        return self.font

    def stylesPatternfill(self,color=None):
        '''
        设置单元格样式
        :param color: 单元格颜色
        :return:
        '''
        self.fill = styles.PatternFill(fill_type='solid',start_color=color)
        return self.fill

    def stylesBorder(self):
        '''
        边框设置
        :return:
        '''
        left = styles.Side(border_style='thin')
        right = styles.Side(border_style='thin')
        top = styles.Side(border_style='thin')
        bottom = styles.Side(border_style='thin')
        self.border = styles.Border(left=left,right=right,top=top,bottom=bottom)
        return self.border

    def stylesAlignment(self):
        '''
        设置字体样式,默认垂直
        :return:
        '''
        self.Alignment = styles.Alignment(horizontal='center')
        return self.Alignment

    def cellDimensionsCol(self,col,width):
        '''
        根据传入参数设置列宽
        :param col: 对应的列的小标 int
        :param width: 列宽的数字  int
        :return:
        '''
        colUpper = self.letter.upperDisponse(col)
        self.ws.column_dimensions[colUpper].width = width       #定义列宽


    def cellDimensionsRow(self,row,rowHeight):
        '''
        根据传入参数设置行高
        :param row: 对应行的数字
        :param rowHeight: 对应行高
        :return:
        '''
        self.ws.row_dimensions[row].height = rowHeight          #定义行高

    def cellStyles(self,row,col,
                   ftSize,ftBold=None,ftColor='FF010701',
                   fillColor=None):
        '''
        设置对应单元格的样式
        :param row: 单元格的行数
        :param col: 单元格列数
        :param ftSize: 字体大小
        :param ftBold: 是否加粗
        :param ftColor: 字体颜色
        :param fillColor: 单元格颜色
        :return:
        '''
        cell = self.ws.cell(row=row,column=col)         #获取单元格属性
        cell.font = self.stylestFout(size=ftSize,bold=ftBold,color=ftColor)
        cell.fill = self.stylesPatternfill(color=fillColor)
        cell.border = self.stylesBorder()
        cell.alignment = self.stylesAlignment()

    def reportInit(self):
        '''
        测试报告首页的生成
        :return:
        '''
        self.ws = self.wb.create_sheet('测试报告')
        self.ws.cell(row=1,column=1,value='测试报告')
        for col in range(1,7):
            self.cellDimensionsCol(col,14.63)       #设置列宽
        for row in range(1,5):
            self.cellDimensionsRow(row,46.5)        #设置行高
        self.cellStyles(1,1,22,ftBold=True,fillColor=excelReport.DoderBlue)
        self.ws.merge_cells(start_row=1,start_column=1,end_column=6,end_row=1)        #合并单元格
        self.ws.cell(row=2,column=1,value='测试项目')
        self.cellStyles(2,1,16,ftBold=True,fillColor=excelReport.white)
        self.ws.cell(row=2,column=3,value='版本')
        self.cellStyles(2,3,16,ftBold=True,fillColor=excelReport.white)
        self.ws.cell(row=2, column=5, value='通过总数')
        self.cellStyles(2,5,16,ftBold=True,fillColor=excelReport.white)
        self.ws.cell(row=3, column=1, value='用例总数')
        self.cellStyles(3,1,16,ftBold=True,fillColor=excelReport.white)
        self.ws.cell(row=3, column=3, value='框架语言')
        self.cellStyles(3,3,16,ftBold=True,fillColor=excelReport.white)
        self.ws.cell(row=3, column=5, value='失败总数')
        self.cellStyles(3,5,16,ftBold=True,fillColor=excelReport.white)
        self.ws.cell(row=4, column=1, value='总耗时')
        self.cellStyles(4, 1, 16, ftBold=True,fillColor=excelReport.white)
        '''对应的标题的值设置'''
        self.ws.cell(row=2, column=2, value='devOps')
        self.cellStyles(2, 2, 16, fillColor=excelReport.white)
        self.ws.cell(row=2, column=4, value='3.7.3')
        self.cellStyles(2, 4, 16, fillColor=excelReport.white)
        self.ws.cell(row=2, column=6, value=str(self.caseTrue))
        self.cellStyles(2, 6, 16, fillColor=excelReport.white)
        self.ws.cell(row=3, column=6, value=str(self.caseFalse))
        self.cellStyles(3, 6, 16, fillColor=excelReport.white)
        self.chartInit()
        self.ws = self.wb["报告详情"]
        for row in range(2,self.ws.max_row):
            color = 'FF010701'
            var = self.ws.cell(row,10).value
            var1 = self.ws.cell(row,1).value
            if var == 'True':
                color = excelReport.LimeGreen
                self.caseTrue += 1
            elif var == 'False':
                color = excelReport.red
                self.caseFalse += 1
            elif '执行excel' in var1:
                color = "FF010701"      #默认使用黑色
                self.cellStyles(row,1,16,fillColor=excelReport.Green)
                self.ws.merge_cells(start_row=row,start_column=1,end_row=row,end_column=11)
                continue
            elif '执行sheet' in var1:
                color = "FF010701"      #默认使用黑色
                self.cellStyles(row,1,14,fillColor=excelReport.LimeGreen)
                self.ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=11)
                continue
            for col in range(1,self.ws.max_column):
                self.cellStyles(row,col,11,ftColor=color,fillColor=excelReport.white)
        self.ws = self.wb["测试报告"]
        self.ws.cell(row=3, column=2, value=str(self.caseFalse + self.caseTrue))
        self.cellStyles(3, 2, 16, fillColor=excelReport.white)
        init = (self.caseFalse + self.caseTrue) / self.caseTrue
        self.ws.cell(row=3, column=4, value=str(round(init)))
        self.cellStyles(3, 4, 16, fillColor=excelReport.white)
        self.wb.save('11.xlsx')

    def chartInit(self):
        '''
        画一个饼图
        :return:
        '''
        self.pie = PieChart()
        lables = Reference(self.ws,min_col=5,min_row=2,max_row=3)
        data = Reference(self.ws,min_col=6,min_row=2,max_row=3)
        self.pie.add_data(data)
        self.pie.set_categories(lables)
        self.pie.title = '通过率对比图'
        self.ws.add_chart(self.pie,'B8')

    def appendInit(self):

        self.ws = self.wb.create_sheet('报告详情',1)
        self.ws.append(['sheet页','ID','index','type','url','parameter','headers','status_code','response','assert','备注'])
        self.cellDimensionsRow(1, 27.75)
        self.cellDimensionsCol(1, 13)
        self.cellDimensionsCol(2, 8.38)
        self.cellDimensionsCol(3, 8.38)
        self.cellDimensionsCol(4, 8.38)
        self.cellDimensionsCol(5, 39)
        self.cellDimensionsCol(6, 33.75)
        self.cellDimensionsCol(7, 16)
        self.cellDimensionsCol(8, 15.38)
        self.cellDimensionsCol(9, 31.25)
        self.cellDimensionsCol(10, 14.13)
        self.cellDimensionsCol(11, 14.13)
        for val in range(1,12):
            self.cellStyles(1, val, 12, ftBold=True, fillColor=excelReport.Orange)




    def excelAppend(self,data):
        if not os.path.exists(os.path.join(os.path.dirname(__file__),'测试报告.xlsx')):
            self.wb = openpyxl.Workbook()
        else:
            self.wb = openpyxl.load_workbook('测试报告.xlsx')
        self.appendInit()       #初始化 测试详情首页
        for value in data:
            self.ws.append(value)
        self.reportInit()
        self.wb.remove(self.wb['Sheet'])
        self.wb.save('测试报告.xlsx')




class letter:

    def __init__(self):
        self.uppercase = []             #大写字母的列表
        self.upperList()

    def upperList(self):
        st = string.ascii_uppercase
        for i in st:
            self.uppercase.append(i)

    def upperDisponse(self,col):
        '''
        获取对应列数的对应字母序号  1对应的是A
        :param col: 传入的列的数字,下标
        :return: 返回组合好的英文字母 大写
        '''
        try:
            val = self.uppercase[col-1]
        except Exception:
            shang = col // 26         #取商值
            yu = col % 26               #取余数
            '''有bug'''
            val = self.uppercase[shang-1] + self.uppercase[yu]
        return val






if __name__ == '__main__':
    # print(excelReport.Gold)
    wb = openpyxl.load_workbook("caogao.xlsx")
    ws = wb["报告模板"]
    ws.cell(1,1).font = styles.Font(color="FF010701")
    wb.save("caogao.xlsx")
    # case = caseExcel()
    # case.status("工作流接口测试用例.xlsx")
    # print(case.excelDic)
    # path = r'C:\Users\Administrator\Desktop\文档\博云BeyondDevOps测试用例.xlsx'
    # wb = xlrd.open_workbook(path)
    # sh = wb.sheet_by_name('POC测试用例')
    # s = sh.cell_value(1,1)
    # print(s)
    # print(sh.nrows)
    # print(sh.cell_value(1,1))
    # a = caseExcel()
    # a.status()
    # for name in a.excelDic:
    #     for case in a.excelDic[name]:
    #         print(case)
    # s = excelReport()
    # s.reportInit()
